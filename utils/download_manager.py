import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
import zipfile

class DownloadManager:
    """Manage downloads of all generated artifacts"""
    
    def __init__(self):
        self.supported_formats = ['pdf', 'xlsx', 'xls', 'zip']
    
    def prepare_download(self, file_id, file_type, format_type, session):
        """
        Prepare file for download based on type and format
        
        Args:
            file_id: ID of the file
            file_type: Type of file (extraction, offer, presentation, mas, ve)
            format_type: Output format (pdf, xlsx, xls, zip)
            session: Flask session object
        
        Returns:
            Path to the file to download
        """
        if format_type not in self.supported_formats:
            raise Exception(f'Unsupported format: {format_type}')
        
        # Get file info
        uploaded_files = session.get('uploaded_files', [])
        file_info = None
        
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                break
        
        if not file_info:
            raise Exception('File not found')
        
        session_id = session['session_id']
        
        # Handle different file types
        if file_type == 'extraction':
            return self.prepare_extraction_download(file_info, format_type, session_id)
        elif file_type == 'offer':
            return self.prepare_offer_download(file_info, format_type, session_id)
        elif file_type == 'presentation':
            return self.prepare_presentation_download(file_info, format_type, session_id)
        elif file_type == 'mas':
            return self.prepare_mas_download(file_info, format_type, session_id)
        elif file_type == 've':
            return self.prepare_ve_download(file_info, format_type, session_id)
        elif file_type == 'all':
            return self.prepare_all_downloads(file_info, session_id)
        else:
            raise Exception(f'Unknown file type: {file_type}')
    
    def prepare_extraction_download(self, file_info, format_type, session_id):
        """Prepare extracted table data for download"""
        if 'extraction_result' not in file_info:
            raise Exception('No extraction data available')
        
        extraction_result = file_info['extraction_result']
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type in ['xlsx', 'xls']:
            return self.create_extraction_excel(extraction_result, output_dir, file_info['id'])
        elif format_type == 'pdf':
            # Return existing extraction output if available
            if 'output_dir' in file_info:
                # Find PDF in output directory
                pdf_files = [f for f in os.listdir(file_info['output_dir']) if f.endswith('.pdf')]
                if pdf_files:
                    return os.path.join(file_info['output_dir'], pdf_files[0])
            raise Exception('PDF extraction not available. Generate offer first.')
        
        raise Exception(f'Format {format_type} not supported for extraction')
    
    def prepare_offer_download(self, file_info, format_type, session_id):
        """Prepare offer for download"""
        if 'costed_data' not in file_info:
            raise Exception('No costed data available. Apply costing first.')
        
        costed_data = file_info['costed_data']
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type in ['xlsx', 'xls']:
            return self.create_offer_excel(costed_data, output_dir, file_info['id'])
        elif format_type == 'pdf':
            # Check if offer PDF was generated
            offer_dir = os.path.join('outputs', session_id, 'offers')
            if os.path.exists(offer_dir):
                pdf_files = [f for f in os.listdir(offer_dir) if f.endswith('.pdf') and file_info['id'] in f]
                if pdf_files:
                    return os.path.join(offer_dir, pdf_files[0])
            raise Exception('Offer PDF not generated yet')
        
        raise Exception(f'Format {format_type} not supported for offers')
    
    def prepare_presentation_download(self, file_info, format_type, session_id):
        """Prepare presentation for download"""
        presentation_dir = os.path.join('outputs', session_id, 'presentations')
        
        if not os.path.exists(presentation_dir):
            raise Exception('Presentation not generated yet')
        
        pdf_files = [f for f in os.listdir(presentation_dir) if f.endswith('.pdf') and file_info['id'] in f]
        if pdf_files:
            return os.path.join(presentation_dir, pdf_files[0])
        
        raise Exception('Presentation file not found')
    
    def prepare_mas_download(self, file_info, format_type, session_id):
        """Prepare MAS for download"""
        mas_dir = os.path.join('outputs', session_id, 'mas')
        
        if not os.path.exists(mas_dir):
            raise Exception('MAS not generated yet')
        
        pdf_files = [f for f in os.listdir(mas_dir) if f.endswith('.pdf') and file_info['id'] in f]
        if pdf_files:
            return os.path.join(mas_dir, pdf_files[0])
        
        raise Exception('MAS file not found')
    
    def prepare_ve_download(self, file_info, format_type, session_id):
        """Prepare value engineering alternatives for download"""
        if 'value_engineering' not in file_info:
            raise Exception('Value engineering not performed yet')
        
        ve_data = file_info['value_engineering']
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        if format_type in ['xlsx', 'xls']:
            return self.create_ve_excel(ve_data, output_dir, file_info['id'])
        
        raise Exception(f'Format {format_type} not supported for value engineering')
    
    def prepare_all_downloads(self, file_info, session_id):
        """Create a ZIP file with all generated documents"""
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        zip_filename = os.path.join(output_dir, f'all_documents_{file_info["id"]}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip')
        
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add extraction data
            if 'extraction_result' in file_info:
                try:
                    excel_file = self.create_extraction_excel(
                        file_info['extraction_result'], 
                        output_dir, 
                        file_info['id']
                    )
                    zipf.write(excel_file, os.path.basename(excel_file))
                except:
                    pass
            
            # Add offer
            if 'costed_data' in file_info:
                try:
                    offer_file = self.create_offer_excel(
                        file_info['costed_data'], 
                        output_dir, 
                        file_info['id']
                    )
                    zipf.write(offer_file, os.path.basename(offer_file))
                except:
                    pass
            
            # Add PDFs from various directories
            for subdir in ['offers', 'presentations', 'mas']:
                dir_path = os.path.join('outputs', session_id, subdir)
                if os.path.exists(dir_path):
                    for filename in os.listdir(dir_path):
                        if file_info['id'] in filename:
                            file_path = os.path.join(dir_path, filename)
                            zipf.write(file_path, f'{subdir}/{filename}')
        
        return zip_filename
    
    def create_extraction_excel(self, extraction_result, output_dir, file_id):
        """Create Excel file from extraction result"""
        filename = os.path.join(output_dir, f'extraction_{file_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each page
        for idx, layout_result in enumerate(extraction_result.get('layoutParsingResults', [])):
            markdown_text = layout_result.get('markdown', {}).get('text', '')
            
            # Extract tables
            tables = self.parse_markdown_tables(markdown_text)
            
            for table_idx, table in enumerate(tables):
                sheet_name = f'Page{idx+1}_Table{table_idx+1}'[:31]  # Excel sheet name limit
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers
                if table['headers']:
                    ws.append(table['headers'])
                    self.style_header_row(ws, 1)
                
                # Add data rows
                for row in table['rows']:
                    row_data = [row.get(h, '') for h in table['headers']]
                    ws.append(row_data)
                
                # Auto-adjust column widths
                self.auto_adjust_columns(ws)
        
        wb.save(filename)
        return filename
    
    def create_offer_excel(self, costed_data, output_dir, file_id):
        """Create Excel file for offer with costing applied"""
        filename = os.path.join(output_dir, f'offer_{file_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Offer'
        
        # Title
        ws.append(['COMMERCIAL OFFER'])
        ws.merge_cells('A1:F1')
        self.style_title_row(ws, 1)
        
        ws.append([])  # Empty row
        
        # Factors applied
        factors = costed_data['factors']
        ws.append(['Costing Factors Applied:'])
        ws.append([f'Net Margin: {factors.get("net_margin", 0)}%'])
        ws.append([f'Freight: {factors.get("freight", 0)}%'])
        ws.append([f'Customs: {factors.get("customs", 0)}%'])
        ws.append([f'Installation: {factors.get("installation", 0)}%'])
        ws.append([f'Exchange Rate: {factors.get("exchange_rate", 1.0)}'])
        ws.append([f'Additional: {factors.get("additional", 0)}%'])
        ws.append([])  # Empty row
        
        # Tables
        for table_idx, table in enumerate(costed_data['tables']):
            ws.append([f'Item List {table_idx + 1}'])
            ws.merge_cells(f'A{ws.max_row}:F{ws.max_row}')
            
            # Headers
            ws.append(table['headers'])
            self.style_header_row(ws, ws.max_row)
            
            # Data rows
            for row in table['rows']:
                row_data = [row.get(h, '') for h in table['headers']]
                ws.append(row_data)
            
            ws.append([])  # Empty row
        
        # Summary
        subtotal = self.calculate_subtotal(costed_data['tables'])
        vat = subtotal * 0.15
        grand_total = subtotal + vat
        
        ws.append(['', '', '', '', 'Subtotal:', subtotal])
        ws.append(['', '', '', '', 'VAT (15%):', vat])
        ws.append(['', '', '', '', 'Grand Total:', grand_total])
        
        self.style_summary_rows(ws, ws.max_row - 2, ws.max_row)
        
        # Auto-adjust columns
        self.auto_adjust_columns(ws)
        
        wb.save(filename)
        return filename
    
    def create_ve_excel(self, ve_data, output_dir, file_id):
        """Create Excel file for value engineering alternatives"""
        filename = os.path.join(output_dir, f've_alternatives_{file_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Alternatives'
        
        # Title
        ws.append([f'VALUE ENGINEERED ALTERNATIVES - {ve_data["budget_option"].upper()}'])
        ws.merge_cells('A1:H1')
        self.style_title_row(ws, 1)
        ws.append([])
        
        # Process alternatives
        for alt_group in ve_data['alternatives']:
            original = alt_group['original_item']
            
            # Original item
            ws.append(['ORIGINAL ITEM'])
            ws.append(['Description:', original.get('description', '')])
            ws.append(['Quantity:', f"{original.get('qty', '')} {original.get('unit', '')}"])
            ws.append(['Unit Rate:', original.get('unit_rate', '')])
            ws.append(['Total:', original.get('total', '')])
            ws.append([])
            
            # Alternatives
            ws.append(['ALTERNATIVES'])
            headers = ['Brand', 'Model', 'Description', 'Unit Rate', 'Total', 'Lead Time']
            ws.append(headers)
            self.style_header_row(ws, ws.max_row)
            
            for alt in alt_group['alternatives']:
                ws.append([
                    alt['brand'],
                    alt['model'],
                    alt['description'],
                    alt['unit_rate'],
                    alt['total'],
                    alt['lead_time']
                ])
            
            ws.append([])
            ws.append([])  # Double space between items
        
        self.auto_adjust_columns(ws)
        
        wb.save(filename)
        return filename
    
    def parse_markdown_tables(self, markdown_text):
        """Parse tables from markdown text"""
        lines = markdown_text.split('\n')
        tables = []
        current_table = {'headers': [], 'rows': []}
        in_table = False
        
        for line in lines:
            if '|' in line:
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                
                if not in_table:
                    # Start of table - headers
                    current_table['headers'] = cells
                    in_table = True
                elif all(c in '-|: ' for c in line):
                    # Separator line - skip
                    continue
                else:
                    # Data row
                    if len(cells) == len(current_table['headers']):
                        row = dict(zip(current_table['headers'], cells))
                        current_table['rows'].append(row)
            else:
                if in_table and current_table['rows']:
                    tables.append(current_table)
                    current_table = {'headers': [], 'rows': []}
                in_table = False
        
        if in_table and current_table['rows']:
            tables.append(current_table)
        
        return tables
    
    def calculate_subtotal(self, tables):
        """Calculate subtotal from tables"""
        import re
        subtotal = 0.0
        
        for table in tables:
            for row in table['rows']:
                for key, value in row.items():
                    if 'total' in key.lower() and '_original' not in key:
                        try:
                            cleaned = re.sub(r'[^\d.-]', '', str(value))
                            num_value = float(cleaned)
                            subtotal += num_value
                        except:
                            pass
        
        return subtotal
    
    def style_header_row(self, ws, row_num):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def style_title_row(self, ws, row_num):
        """Apply styling to title row"""
        title_fill = PatternFill(start_color='764BA2', end_color='764BA2', fill_type='solid')
        title_font = Font(bold=True, size=16, color='FFFFFF')
        
        for cell in ws[row_num]:
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def style_summary_rows(self, ws, start_row, end_row):
        """Apply styling to summary rows"""
        bold_font = Font(bold=True)
        
        for row_num in range(start_row, end_row + 1):
            for cell in ws[row_num]:
                if cell.value:
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='right')
    
    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
